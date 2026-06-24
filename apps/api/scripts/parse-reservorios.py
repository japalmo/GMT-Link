import os
import re
import json
import datetime
import openpyxl

KML_DIR = r"C:\Users\juana\Downloads\poligonos reservorios"
EXCEL_PATH = r"C:\Users\juana\Downloads\VOLUMEN 3D DE RESERVORIOS DIA 18-06-2026.xlsx"
OUT_PATH = r"C:\Users\juana\GMT Link\apps\api\prisma\data-reservorios.json"

# Mapeo de reservorios a sus hojas correspondientes
SHEETS_MAP = {
    "R1": ["R1"],
    "R2": ["R2", "R2-"], # R2 se divide en dos hojas cronológicas
    "R3": ["R3"],
    "R4": ["R4"],
    "R5": ["R5"],
    "R6": ["R6 NEW"],
    "R7": ["R7"],
    "R8": ["R8"],
    "R9": ["R9 NEW"],
    "R10": ["R10 New"]
}

# Metadatos base y capacidades extraídas de la pestaña general de Excel
CAPACITIES = {
    "R1": 4835.0,
    "R2": 6311.0,
    "R3": 7648.0,
    "R4": 8763.0,
    "R5": 6931.0,
    "R6": 6010.0,
    "R7": 6418.0,
    "R8": 6745.0,
    "R9": 6256.0,
    "R10": 6325.0
}

def parse_kml(code):
    filepath = os.path.join(KML_DIR, f"{code}.kml")
    if not os.path.exists(filepath):
        print(f"KML para {code} no encontrado.")
        return []
    
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
        match = re.search(r"<coordinates>(.*?)</coordinates>", content, re.DOTALL)
        if match:
            coords_str = match.group(1).strip()
            # KML es lon,lat,alt. Nosotros queremos [lat, lon]
            coords = []
            for pt in coords_str.split():
                parts = pt.split(",")
                if len(parts) >= 2:
                    coords.append([float(parts[1]), float(parts[0])])
            return coords
    return []

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", ".").strip())
    except:
        return None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    output = {}

    for i in range(1, 11):
        code = f"R{i}"
        name = f"Reservorio {i}"
        
        # 1. Parse KML
        polygon = parse_kml(code)
        
        # 2. Parse Excel measurements
        raw_rows = []
        sheets = SHEETS_MAP[code]
        
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                print(f"Hoja {sheet_name} no encontrada.")
                continue
            
            ws = wb[sheet_name]
            # Iterar sobre las filas saltando cabeceras
            for row in list(ws.iter_rows(min_row=3)):
                date_val = row[0].value
                if date_val is None:
                    continue
                
                # Intentar parsear fecha
                dt = None
                if isinstance(date_val, datetime.datetime):
                    dt = date_val
                elif isinstance(date_val, str):
                    try:
                        dt = datetime.datetime.strptime(date_val.strip().split()[0], "%Y-%m-%d")
                    except:
                        try:
                            dt = datetime.datetime.strptime(date_val.strip().split()[0], "%d/%m/%Y")
                        except:
                            continue
                
                if not dt:
                    continue
                
                # Extraer las columnas específicas
                borde_libre_cm = safe_float(row[2].value)
                altura_salmuera_cm = safe_float(row[3].value)
                altura_sal_cm = safe_float(row[4].value)
                cota_espejo = safe_float(row[5].value)
                cota_sal = safe_float(row[6].value)
                area_espejo = safe_float(row[7].value)
                perimetro = safe_float(row[8].value)
                vol_libre = safe_float(row[10].value)
                vol_sal = safe_float(row[17].value)
                
                # Evitar registros vacíos o ruidosos
                if vol_libre is None and cota_espejo is None:
                    continue
                
                # Normalización de datos
                borde_libre = (borde_libre_cm / 100.0) if borde_libre_cm is not None else None
                altura_salmuera = (altura_salmuera_cm / 100.0) if altura_salmuera_cm is not None else None
                altura_sal = (altura_sal_cm / 100.0) if altura_sal_cm is not None else None
                vol_ocluida = (vol_sal * 0.20) if vol_sal is not None else 0.0
                vol_total = (vol_libre + vol_ocluida) if vol_libre is not None else 0.0
                
                raw_rows.append({
                    "date": dt.strftime("%Y-%m-%d"),
                    "timestamp": dt.timestamp(),
                    "borde_libre": borde_libre,
                    "altura_salmuera": altura_salmuera,
                    "altura_sal": altura_sal,
                    "cota_espejo": cota_espejo,
                    "cota_sal": cota_sal,
                    "area_espejo": area_espejo,
                    "perimetro": perimetro,
                    "vol_salmuera_libre": vol_libre,
                    "vol_sal": vol_sal,
                    "vol_salmuera_ocluida": vol_ocluida,
                    "vol_total_salmuera": vol_total
                })
        
        # Eliminar duplicados de fecha y ordenar cronológicamente
        unique_rows = {}
        for r in raw_rows:
            unique_rows[r["date"]] = r
        
        sorted_rows = sorted(unique_rows.values(), key=lambda x: x["timestamp"])
        
        # Downsampling si hay demasiadas filas (> 100) para asegurar fluidez en gráficos
        if len(sorted_rows) > 100:
            step = len(sorted_rows) / 100.0
            downsampled = []
            for j in range(100):
                idx = int(j * step)
                if idx < len(sorted_rows):
                    downsampled.append(sorted_rows[idx])
            # Asegurar que el último registro real (el del día 18-06-2026) siempre esté incluido
            if sorted_rows[-1]["date"] != downsampled[-1]["date"]:
                downsampled[-1] = sorted_rows[-1]
            sorted_rows = downsampled
            
        # Calcular cotas estáticas
        cotas_espejo = [r["cota_espejo"] for r in sorted_rows if r["cota_espejo"] is not None]
        max_cota = max(cotas_espejo) if cotas_espejo else 2302.0
        min_cota = min(cotas_espejo) if cotas_espejo else 2300.5
        
        # Límite máximo de capacidad en el Excel
        cap_max = CAPACITIES[code]
        
        metadata = {
            "cota_fondo": round(min_cota - 1.5, 3),
            "cota_segura": round(max_cota - 0.2, 3),
            "cota_lamina_critica": round(max_cota + 0.1, 3),
            "limits": {
                "effective_area": round(cap_max / 1.2, 2), # Área Operacional Efectiva
                "safe_capacity": cap_max,                  # Capacidad Operativa de Seguridad
                "max_nominal_capacity": round(cap_max * 1.15, 2) # Capacidad Hidráulica Máxima
            }
        }
        
        output[code] = {
            "code": code,
            "name": name,
            "polygon": polygon,
            "metadata": metadata,
            "measurements": sorted_rows
        }
        
        print(f"Reservorio {code}: {len(sorted_rows)} registros procesados.")

    # Guardar a archivo JSON
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
        
    print(f"\nDatos guardados exitosamente en {OUT_PATH}")

if __name__ == "__main__":
    main()
